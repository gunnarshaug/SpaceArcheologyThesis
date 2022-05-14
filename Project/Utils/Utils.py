import itertools
import shutil

import openpyxl
import os
import json

import pandas

datasets_path = os.path.join(os.getcwd(), os.pardir, "Datasets")
json_file = os.path.join(datasets_path, "http", "asia.json")
country = "Uzbekistan"
L = 0.003

def flatten(list_of_lists):
    if len(list_of_lists) == 0:
        return list_of_lists
    if isinstance(list_of_lists[0], list):
        return flatten(list_of_lists[0]) + flatten(list_of_lists[1:])
    return list_of_lists[:1] + flatten(list_of_lists[1:])

def format(num):
    return round(num, 5)


def writeToTarget(data, target):
    workbook = openpyxl.load_workbook(target)
    sheet = workbook.active
    sheet.delete_rows(2, sheet.max_row)

    for data_entry in data:
        lat, long, id = data_entry

        longitude_top_left = long - L
        longitude_bottom_right = long + L
        latitude_bottom_right = lat - L
        latitude_top_left = lat + L
        sheet.append([
            format(longitude_top_left),
            format(latitude_top_left),
            format(longitude_bottom_right),
            format(latitude_bottom_right),
            country,
            id
        ])
    workbook.save(target_file)


def create_list_from_text_file(source, target):
    file = open(source, "r")
    content = file.read()
    lines = content.split("\n")
    data = []
    for line in lines:
        if not line.strip():
            continue
        lat, long, id = line.split(",")
        data.append((float(lat), float(long), id))

    writeToTarget(data, target)


def load_json(kind=None, reject=[]):
    entries = []
    data = open(json_file)
    content = json.load(data)
    types = []
    for entry in content.get('features'):
        props = entry.get('properties')

        # if props.get('id') == 2438:
        #     print(entry)

        if props.get('isvisible') == 0:
            continue

        if not props.get('kind') in types:
            types.append(props.get('kind'))

        if props.get('kind') in reject:
            continue

        if kind is None:
            entries.append(entry)
            continue

        if props.get('kind') == kind:
            entries.append(entry)
            continue

    return [(entry.get('geometry').get('coordinates'), entry.get('properties').get('id')) for entry in entries]


def create_source_file(coords, file_path):
    with open(file_path, 'w') as file:
        for (coord, id) in coords:
            long, lang = coord
            file.write(",".join([str(lang), str(long), str(id)]))
            file.write("\n")

def move_images(source, target):
    file_paths = get_filepaths(source)

    for file in file_paths:
        filename = os.path.basename(file)
        filename = os.path.join(target, filename)
        shutil.copyfile(file, os.path.join(target, filename))

    #print(file_paths)




def get_filepaths(source):
    content = os.listdir(source)
    paths = [os.path.join(source, folder) for folder in content]

    files = [file for file in paths if os.path.isfile(file)]

    folders = [folder for folder in paths if os.path.isdir(folder)]

    for folder in folders:
        file_paths = get_filepaths(folder)
        files.append(file_paths)

    return flatten(files)

def cleanup_folder():
    target_path = os.path.join(datasets_path, "Data")
    classification_file = os.path.join(datasets_path, 'classification.csv')

    entries = pandas.read_csv(classification_file)
    file_names = []

    for i in range(len(entries)):
        entry = entries.iloc[i]
        file_names.append(entry.filename)

    for file in os.listdir(target_path):
        if file not in file_names:
            file_path = os.path.join(target_path, file)
            os.remove(file_path)



if __name__ == "__main__":
    ###################
    # Create forts file
    ###################



    #target_file = os.path.join(datasets_path, "Negative_Fortress.xlsx")
    target_file = os.path.join(datasets_path, "asia.xlsx")
    source_file = os.path.join(datasets_path, "coords.txt")

    # target_file = os.path.join(datasets_path, "worldcities.xlsx")
    #
    # dataset_list = []
    # workbook = openpyxl.load_workbook(target_file)
    # sheet = workbook.active
    # count = 200
    # i = 0
    #
    # for row in sheet.iter_rows(min_row=2):
    #     cell_values = tuple([cell.value for cell in row])
    #     _, _, lat, long, _, _, _, _, _, population, id = cell_values
    #     dataset_list.append((lat, long, id, population))
    #     i += 1
    #
    # with open(source_file, 'w') as file:
    #     for (lat, long, id, population) in dataset_list:
    #         if population is None:
    #             continue
    #         if int(population) > 200000 or int(population) < 7000:
    #             continue
    #
    #         if int(lat) < 30 or int(lat) > 41:
    #             continue
    #         if int(long) < 40 or int(long) > 75:
    #             continue
    #
    #         file.write(",".join([str(lat), str(long), str(id)]))
    #         file.write("\n")
    # ['city', 'vicus', 'event', 'fort', 'altar', 'temple', 'observation', 'graves', 'building', 'theater', 'object', 'industry', 'rural', 'aquaduct', 'mansio']
    #forts_coords = load_json(reject=['observation', 'object', 'museum', 'vicus', 'city', 'rural', 'graves'])
    forts_coords = load_json(kind="fort")
    create_source_file(forts_coords, source_file)
    create_list_from_text_file(source_file, target_file)

    ###################
    # Create non forts file
    ###################
    #target_file = os.path.join(datasets_path, "Negative_Fortress.xlsx")
    #source_file = os.path.join(datasets_path, "rest_coords.txt")

    # rest_coords = load_json(None, "fort")
    # create_source_file(rest_coords, source_file)
    # create_list_from_text_file(source_file, target_file)

    ###################
    # Move images around
    ###################
    # source = os.path.join(datasets_path, "Negative_Fortresses")
    # target = os.path.join(datasets_path, "Data")
    # move_images(source, target)

    ###################
    # Cleanup
    ###################
    #cleanup_folder()




